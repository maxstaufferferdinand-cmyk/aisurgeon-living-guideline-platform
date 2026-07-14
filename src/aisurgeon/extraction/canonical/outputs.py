"""Non-overwriting canonical JSONL and human-review workbook outputs."""

import json
import os
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from aisurgeon.extraction.canonical.models import ReviewFinding


def _exclusive_write(path: Path, content: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor = os.open(path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
    with os.fdopen(descriptor, "wb") as stream:
        stream.write(content)


def write_jsonl(path: Path, records: Iterable[Any]) -> None:
    lines = []
    for record in records:
        value = record.model_dump(mode="json") if hasattr(record, "model_dump") else record
        lines.append(json.dumps(value, ensure_ascii=False, separators=(",", ":")))
    _exclusive_write(path, (("\n".join(lines) + "\n") if lines else "").encode())


def write_json(path: Path, value: Any) -> None:
    if hasattr(value, "model_dump"):
        value = value.model_dump(mode="json")
    _exclusive_write(path, (json.dumps(value, ensure_ascii=False, indent=2) + "\n").encode())


def write_review_workbook(path: Path, findings: list[ReviewFinding]) -> None:
    if path.exists():
        raise FileExistsError("Review-Workbook existiert bereits; nichts überschrieben.")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "review_findings"
    headers = list(ReviewFinding.model_fields)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for finding in findings:
        value = finding.model_dump(mode="json")
        sheet.append([value.get(name) for name in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).coordinate}"
    for index, header in enumerate(headers, start=1):
        width = min(60, max(12, len(header) + 2))
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    workbook.save(path)
    if os.name == "posix":
        path.chmod(0o600)


CANONICAL_OUTPUTS = (
    "document_map.validated.json",
    "formal_items.jsonl",
    "recommendations.jsonl",
    "statements.jsonl",
    "comments.jsonl",
    "references.jsonl",
    "tables.jsonl",
    "algorithms.jsonl",
    "decision_trees.jsonl",
    "clinical_context_blocks.jsonl",
    "unresolved_links.jsonl",
    "review_findings.jsonl",
    "review_findings.xlsx",
    "extraction_summary.json",
    "extraction_manifest.json",
)
